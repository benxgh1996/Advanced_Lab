import openpyxl
import math

def get_data(wb_name, num_trials):
	wb = openpyxl.load_workbook(wb_name, read_only=True, data_only=True)
	ws = wb.get_sheet_by_name("charge_mass")
	# data is a list of [acceleration_voltage, coil_current, ring_radius]
	data = []
	for r in range(2, 2 + num_trials):
		voltage = float(ws.cell(row=r, column=1).value)
		coil_current = float(ws.cell(row=r, column=2).value)
		ring_radius = float(ws.cell(row=r, column=3).value) / 2.0 / 100.0
		data.append([voltage, coil_current, ring_radius])
	wb.close()
	return data


def  get_mag_field(coil_current):
	MIU_0 = 4 * math.pi * 10 ** (-7)
	NUM_TURNS_COIL = 130
	COIL_RADIUS = 0.15
	mag_field = (8 * MIU_0 * NUM_TURNS_COIL * coil_current) / (5 * math.sqrt(5) * COIL_RADIUS)
	return mag_field


def get_charge_mass_ratio(voltage, coil_current, ring_radius):
	mag_field = get_mag_field(coil_current)
	speed = 2 * voltage / (ring_radius * mag_field)
	charge_mass_ratio = speed / (mag_field * ring_radius)
	return charge_mass_ratio


def get_results(data):
	# results is a list of [voltage, coil_current, charge_mass_ratio]
	results = []
	for point in data:
		voltage = point[0]
		coil_current = point[1]
		ring_radius = point[2]
		charge_mass_ratio = get_charge_mass_ratio(voltage, coil_current, ring_radius)
		results.append([voltage, coil_current, charge_mass_ratio])
	return results


def export_results(wb_name, results):
	wb = openpyxl.load_workbook(wb_name)
	ws = wb.get_sheet_by_name("charge_mass")
	STARTING_ROW = 2
	r = STARTING_ROW

	for point in results:
		ws.cell(row=r, column=1).value = point[0]
		ws.cell(row=r, column=2).value = point[1]
		ws.cell(row=r, column=3).value = point[2]
		r += 1
	wb.save(wb_name)


def get_ave_charge_mass_ratio(results):
	s = 0
	for point in results:
		s += point[2]
	return s / float(len(results))


data = get_data("data.xlsx", 13)
results = get_results(data)
export_results("results.xlsx", results)
# print len(results)
# print results
# print get_ave_charge_mass_ratio(results)


