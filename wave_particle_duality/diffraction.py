import openpyxl
import math
import sys

path_length = 136.4
bulb_radius = 128.4 / 2.0

def get_data(wb_name, num_trials, diameter_type):
	wb = openpyxl.load_workbook(wb_name, read_only=True, data_only=True)
	ws = wb.get_sheet_by_name("diffraction")
	# data is a list of [acceleration_voltage, small_radius, large_radius]
	data = []
	
	if diameter_type == "central":
		col_diameter_small = 14
		col_diameter_large = 15
	elif diameter_type == "inner":
		col_diameter_small = 10
		col_diameter_large = 12
	elif diameter_type == "outer":
		col_diameter_small = 11
		col_diameter_large = 13
	else:
		print "Wrong type of diameter."
		sys.exit()

	for r in range(2, 2 + num_trials):
		voltage = float(ws.cell(row=r, column=1).value)
		small_diameter = float(ws.cell(row=r, column=col_diameter_small).value)
		large_diameter = float(ws.cell(row=r, column=col_diameter_large).value)
		data.append([voltage, small_diameter/2.0, large_diameter/2.0])
	wb.close()
	return data


def get_sin_theta(radius):
	alpha = radius / bulb_radius
	sin_alpha = math.sin(alpha)
	cos_alpha = math.cos(alpha)
	tan_2_theta = sin_alpha / (path_length / bulb_radius + cos_alpha - 1)
	two_theta = math.atan(tan_2_theta)
	if two_theta < 0:
		two_theta += math.pi
	theta = two_theta / 2.0
	sin_theta = math.sin(theta)
	return sin_theta
	

def get_diffraction_angles(data):
	# results is list of [1/sqrt(voltage), sin_theta_small, sin_theta_large]
	results = []
	for point in data:
		voltage = 1.0 / math.sqrt(point[0])
		small_radius = point[1]
		large_radius = point[2]
		sin_theta_small = get_sin_theta(small_radius)
		sin_theta_large = get_sin_theta(large_radius)
		results.append([voltage, sin_theta_small, sin_theta_large])
	return results


def export_results(wb_name, results, diameter_type):
	wb = openpyxl.load_workbook(wb_name)
	ws = wb.get_sheet_by_name("diffraction")
	STARTING_ROW = 3
	r = STARTING_ROW
	
	if diameter_type == "central":
		col_sin_small = 2
		col_sin_large = 3
	elif diameter_type == "inner":
		col_sin_small = 4
		col_sin_large = 5
	elif diameter_type == "outer":
		col_sin_small = 6
		col_sin_large = 7
	else:
		sys.exit()
		
	for point in results:
		ws.cell(row=r, column=1).value = point[0]
		ws.cell(row=r, column=col_sin_small).value = point[1]
		ws.cell(row=r, column=col_sin_large).value = point[2]
		r += 1
	wb.save(wb_name)

DIAMETER_TYPE = "outer"
data = get_data("data.xlsx", 15, DIAMETER_TYPE)
results = get_diffraction_angles(data)

print len(results)
print results
# export_results("results.xlsx", results, DIAMETER_TYPE)

