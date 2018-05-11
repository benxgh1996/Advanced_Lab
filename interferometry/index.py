import openpyxl
import numpy
import math

RECORDING_INTERVAL = 0.001

def true_pressure(pressure_reading):
	VACCUM_PRESSURE_READING = -8.335
	AIR_PRESSURE_READING = 0.010
	VACCUM_PRESSURE = 0
	AIR_PRESSURE = 1

	pressure = (pressure_reading - VACCUM_PRESSURE_READING) / (AIR_PRESSURE_READING - VACCUM_PRESSURE_READING) \
			   * AIR_PRESSURE
	return pressure


def read_wb(wb_name):
	wb = openpyxl.load_workbook(wb_name, read_only=True)
	ws = wb.active
	wb_min_row = ws.min_row + 1
	wb_max_row = ws.max_row
	wb_min_col = ws.min_column
	wb_max_col = ws.max_column
	data = []
	for row in ws.iter_rows(min_row=wb_min_row, max_row=wb_max_row, min_col=wb_min_col, max_col=wb_max_col):
		tmp_data = []
		for cell in row:
			tmp_data.append(float(cell.value))
		data.append(tmp_data)
	wb.close()
	return data

def adjust_pressure(data):
	for single_data in data:
		single_data[2] = true_pressure(single_data[2])


def print_sample_data(data):
	length = len(data)
	print "length: ", length
	print "1: ", data[0]
	print "2: ", data[1]
	print str(length / 2) + ": ", data[length / 2 - 1]
	print str(length / 2 + 1) + ": ", data[length / 2]
	print str(length - 1) + ": ", data[length - 2]
	print str(length) + ": ", data[length - 1]
	

def is_data_group_ambiguous(data_group):
	isAmbiguous = False
	for i in range(len(data_group) - 1):
		if data_group[i][2] == data_group[i + 1][2]:
			print("Ambiguity at: ", data_group[i], data_group[i + 1])
			isAmbiguous = True
	return isAmbiguous

def get_data_group(data):
	# define a data point to be one row of data [time, intensity, pressure] in our original excel input.

	# data_group is a list that will hold groups of data points as single elements. Each group will be a 
	# summation of several data points in terms of their intensity values.
	data_group = []
	# i documents the index of the currently iterated data point in the data list. we used i modulo SEQ_LENGTH to
	# determine whether we have reached an end of a group.
	i = 0
	s = 0
	group_number = 0
	for single_data in data:
		s += single_data[1]
		# when this if statement is true, it means we reach the last data point in a group.
		if i % SEQ_LENGTH == SEQ_LENGTH - 1:
			# data_group will have elements in form of [group number, data list index of the first data point
			# in a group, group sum]
			data_group.append([group_number, group_number * SEQ_LENGTH, s])
			group_number += 1
			# refreshes summation for a group.
			s = 0
		i += 1
	return data_group


def get_minimums(data_group):
	# now we have obtained data_group.
	# minimums will hold the data_group elements that prove to have local maximum intensities. 
	minimums = []
	# we only interated less the first and the last elements of the [data_group]
	for i in range(1, len(data_group) - 1):
		if data_group[i][2] < -0.8 * SEQ_LENGTH:
			if data_group[i][2] <= data_group[i - 1][2] and data_group[i][2] <= data_group[i + 1][2]:
				minimums.append(data_group[i])
				if data_group[i][2] == data_group[i - 1][2] or data_group[i][2] == data_group[i + 1][2]:
					print "ambiguity at: ", data_group[i - 1], data_group[i], data_group[i + 1]
	return minimums


def get_sig_pressures(minimums):
	# now we have obtained minimums.
	# now we want to find the data list index of the local maximum intensity points. 
	# we assume the middle point in each local-min data_group element is the local maximum intensity points. 
	sig_pressures = []
	for group in minimums:
		sig_index = group[1] + SEQ_LENGTH / 2
		# elements of sig_pressures will have form [time, pressure]
		sig_pressures.append([data[sig_index][0], data[sig_index][2]])
	return sig_pressures


def getPressureSTDinGroup(minimums):
	# We want to obtain the standard deviations of pressures in the groups of minimum intensities.
	# Then we want to average these standard deviations across all these minimum groups.
	groups_of_group_pressure = []
	for group in minimums:
		first_index = group[1]
		last_index = group[1] + SEQ_LENGTH - 1
		group_pressures = []
		for i in range(first_index, last_index + 1):
			group_pressures.append(data[i][2])
		groups_of_group_pressure.append(group_pressures)
	print "Length of groups_of_group_pressures: ", len(groups_of_group_pressure)
	sum_std_pressures = 0
	for group in groups_of_group_pressure:
		sum_std_pressures += numpy.std(group)
	print "Average STD of pressures in a group: ", sum_std_pressures / float(len(groups_of_group_pressure))
	
	
def get_max_intensity_points(data):
	data_group = get_data_group(data)
	minimums = get_minimums(data_group)
	get_sig_pressures(minimums)
	

def get_pressure_intervals(sig_pressures):
	p_intervals = []
	for i in range(1, len(sig_pressures)):
		p_intervals.append(sig_pressures[i][1] - sig_pressures[i - 1][1])
	return p_intervals


def write_sig_pressures(sig_pressures, wb_name, trial_num):
	wb = openpyxl.load_workbook(wb_name)
	ws = wb.active
	STARTING_ROW = 3
	
	r = STARTING_ROW
	c = trial_num * 2 - 1
	for point in sig_pressures:
		ws.cell(row=r, column=c).value = point[0]
		ws.cell(row=r, column=c+1).value = point[1]
		r += 1
	wb.save(wb_name)
	
def testSTDPressures(data, first_index):
	pressures = []
	for i in range(first_index, first_index + SEQ_LENGTH):
		pressures.append(data[i][2])
	print "TestedPressureSTD at index ", first_index, "is: ", numpy.std(pressures)


def getSTD(sample_list, standard):
	a_list = list(sample_list)
	for i in range(len(a_list)):
		a_list[i] -= standard
		a_list[i] **= 2
	return math.sqrt(sum(a_list)/float(len(a_list)))

def getDeltaPressureSTD(wb_name, standard):
	wb = openpyxl.load_workbook(wb_name, read_only=True)
	ws = wb.get_sheet_by_name("delta_p")
	p_changes = []
	for r in range(2, 40):
		p_changes.append(ws.cell(row=r, column=2).value)
	wb.close()
	print "Number of all Delta Ps: ", len(p_changes)
	print "Delta Ps are: ", p_changes
	print "Delta P STD for all trials is: ", getSTD(p_changes, standard)
	print "Reference STD is: ", numpy.std(p_changes)


# SEQ_LENGTH_VALUES: (air1, 50), (air2, 40), (air3, 100), (air4, 75), (air5, 100), (air6, 100)
# (helium_1, 100), (helium_2, 100), (helium_3, 100), (helium_4, 100), (helium_5, 100), (helium_6, 100)
# (helium_f1, 100), (helium_f2, 100), (helium_f3, 100), (helium_f4, 100)
SEQ_LENGTH = 100
data = read_wb("helium_4.xlsx")
adjust_pressure(data)
# print_sample_data(data)
data_group = get_data_group(data)
minimums = get_minimums(data_group)

# print "Length of minimums: ", len(minimums)
# getPressureSTDinGroup(minimums)
# testSTDPressures(data, 4000)

sig_pressures = get_sig_pressures(minimums)
getDeltaPressureSTD("significant_pressures_helium.xlsx", 0.066651)

# write_sig_pressures(sig_pressures, "significant_pressures_helium.xlsx", 10)


# print(len(sig_pressures))
# for point in sig_pressures:
# 	print point
# print ""
# print get_pressure_intervals(sig_pressures)


